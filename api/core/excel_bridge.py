# api/core/excel_bridge.py
import sys
import json
import difflib
import copy
import io
import ast
import re
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# ==========================================
# 1. منظومة سمات التصميم (Design System Tokens)
# ==========================================

THEME_PRESETS = {
    "etheer_gold": {
        "header_fill": PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid"),
        "header_font": Font(name="Calibri", size=11, bold=True, color="D4AF37"),
        "accent_fill": PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid"),
        "accent_font": Font(name="Calibri", size=11, bold=True, color="1A1A1A"),
        "zebra_fill": PatternFill(start_color="F9F8F3", end_color="F9F8F3", fill_type="solid"),
    },
    "corporate_blue": {
        "header_fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "accent_fill": PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid"),
        "accent_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "zebra_fill": PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid"),
    },
    "emerald_finance": {
        "header_fill": PatternFill(start_color="004D40", end_color="004D40", fill_type="solid"),
        "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "accent_fill": PatternFill(start_color="00796B", end_color="00796B", fill_type="solid"),
        "accent_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "zebra_fill": PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"),
    },
    "minimal_dark": {
        "header_fill": PatternFill(start_color="212529", end_color="212529", fill_type="solid"),
        "header_font": Font(name="Calibri", size=11, bold=True, color="F8F9FA"),
        "accent_fill": PatternFill(start_color="495057", end_color="495057", fill_type="solid"),
        "accent_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "zebra_fill": PatternFill(start_color="F1F3F5", end_color="F1F3F5", fill_type="solid"),
    }
}

# ==========================================
# 2. الفحص الساكن للأمان (AST Security Guard)
# ==========================================

class SecurityVisitor(ast.NodeVisitor):
    FORBIDDEN_CALLS = {'eval', 'exec', 'open', '__import__', 'compile', 'input'}
    FORBIDDEN_MODULES = {'os', 'sys', 'subprocess', 'shutil', 'socket', 'requests', 'urllib'}

    def __init__(self):
        self.violations = []

    def visit_Import(self, node):
        for alias in node.names:
            if alias.name.split('.')[0] in self.FORBIDDEN_MODULES:
                self.violations.append(f"استيراد مكتبة محظورة: '{alias.name}'")
        self.generic_visit(node)

    def visit_ImportFrom(self, node):
        if node.module and node.module.split('.')[0] in self.FORBIDDEN_MODULES:
            self.violations.append(f"استيراد من مكتبة محظورة: '{node.module}'")
        self.generic_visit(node)

    def visit_Call(self, node):
        if isinstance(node.func, ast.Name) and node.func.id in self.FORBIDDEN_CALLS:
            self.violations.append(f"استدعاء دالة نظام خطيرة: '{node.func.id}'")
        self.generic_visit(node)

def validate_python_code_security(code_str):
    """تحليل شجرة الإعراب المجردة للكود لضمان الخلو من أي عمليات ماسة بنظام التشغيل"""
    try:
        tree = ast.parse(code_str)
        visitor = SecurityVisitor()
        visitor.visit(tree)
        if visitor.violations:
            return False, f"رفض أمني: {', '.join(visitor.violations)}"
        return True, ""
    except Exception as e:
        return False, f"خطأ بنيوي في صياغة الكود: {str(e)}"

# ==========================================
# 3. الأدوات المساعدة وحماية الهيكل والتطبيع
# ==========================================

def normalize_arabic_text(text):
    """توحيد النصوص العربية لتجاوز فروقات الهمزات، التاءات، والمسافات الخفية"""
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = re.sub(r'[\r\n\t]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    
    trans = str.maketrans({
        'أ': 'ا', 'إ': 'ا', 'آ': 'ا',
        'ة': 'ه', 'ى': 'ي',
        '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
        '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'
    })
    return text.translate(trans)

def get_effective_cell_value(ws, row, col):
    """قراءة قيمة الخلية حتى لو كانت واقعة ضمن نطاق مدمج (Merged Range)"""
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val
    
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None

def copy_cell_style(source_cell, target_cell):
    """استنساخ الأنماط والتنسيقات لضمان التطابق التام"""
    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=copy.copy(source_cell.font.color)
            )
        if source_cell.fill and source_cell.fill.fill_type:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=copy.copy(source_cell.fill.start_color),
                end_color=copy.copy(source_cell.fill.end_color)
            )
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
        if source_cell.border:
            target_cell.border = Border(
                left=copy.copy(source_cell.border.left),
                right=copy.copy(source_cell.border.right),
                top=copy.copy(source_cell.border.top),
                bottom=copy.copy(source_cell.border.bottom)
            )
        target_cell.number_format = source_cell.number_format

def find_real_header_row(ws, max_scan_rows=10):
    """كشف ذكي لصف العناوين التجاري وتجاوز العنوان الرئيسي المدمج"""
    best_row = 1
    max_score = -1

    for r in range(1, min(max_scan_rows + 1, ws.max_row + 1)):
        row_merged = [m for m in ws.merged_cells.ranges if m.min_row <= r <= m.max_row]
        if any((m.max_col - m.min_col) > 3 for m in row_merged):
            continue

        row_vals = [
            str(get_effective_cell_value(ws, r, c)).strip()
            for c in range(1, ws.max_column + 1)
            if get_effective_cell_value(ws, r, c) is not None
        ]
        score = len(set(row_vals))
        if score > max_score:
            max_score = score
            best_row = r

    return best_row

def find_header_column(ws, header_input, search_rows=10):
    """محرك كشف الأعمدة المتقدم (يدعم الحروف المباشرة، الأرقام، التطبيع العربي، والاحتواء)"""
    if not header_input:
        return None, None

    header_str = str(header_input).strip()

    # 1. التحديد المباشر عبر رقم العمود أو حرفه (مثل "C" أو "3")
    if header_str.isdigit():
        col_idx = int(header_str)
        if 1 <= col_idx <= ws.max_column:
            return find_real_header_row(ws), col_idx
    elif re.match(r'^[A-Za-z]{1,3}$', header_str):
        try:
            col_idx = column_index_from_string(header_str.upper())
            if 1 <= col_idx <= ws.max_column:
                return find_real_header_row(ws), col_idx
        except ValueError:
            pass

    target_norm = normalize_arabic_text(header_str)

    # 2. التطبيع والتطابق التام
    for r in range(1, min(search_rows + 1, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = get_effective_cell_value(ws, r, c)
            if val is not None and normalize_arabic_text(val) == target_norm:
                return r, c

    # 3. الاحتواء الجزئي المرن
    for r in range(1, min(search_rows + 1, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = get_effective_cell_value(ws, r, c)
            if val is not None:
                val_norm = normalize_arabic_text(val)
                if target_norm in val_norm or val_norm in target_norm:
                    return r, c

    return None, None

def find_sheet_flexible(wb, target_name):
    """مطابقة أوراق العمل بمرونة تجانس الاختلافات اللفظية والرموز"""
    if not target_name:
        return None
        
    available = wb.sheetnames
    if target_name in available:
        return target_name
    
    clean_target = str(target_name).replace("_", " ").replace("-", " ").strip().lower()
    clean_available = [s.replace("_", " ").replace("-", " ").strip().lower() for s in available]
    
    if clean_target in clean_available:
        return available[clean_available.index(clean_target)]
    
    matches = difflib.get_close_matches(clean_target, clean_available, n=1, cutoff=0.65)
    if matches:
        return available[clean_available.index(matches[0])]
    
    return None

def get_sheet(wb, op):
    """تحديد الشيت المستهدف بصورة ديناميكية"""
    sheet_name = op.get("sheet") or op.get("sheet_name")
    if sheet_name:
        matched = find_sheet_flexible(wb, sheet_name)
        if matched:
            return wb[matched]
        raise ValueError(f"الشيت المطلوب غير موجود ولا يوجد اسم مقارب له: {sheet_name}")
    return wb.active

def scan_file_health(ws):
    """فحص سلامة المعادلات وتحديد الخلل المكسور (#REF! / #VALUE!)"""
    broken = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.startswith("="):
                if any(err in val for err in ["#REF!", "#VALUE!", "#NAME?"]):
                    broken.append(f"{get_column_letter(c)}{r}: {val}")
    return broken

# ==========================================
# 4. معالجة التنسيقات والضبط الذكي والسمات
# ==========================================

def op_autofit_columns(wb, ws, op, log):
    """ملاءمة عرض كافة الأعمدة آلياً لمنع اقتطاع المحتوى"""
    padding = op.get("padding", 4)
    max_cap = op.get("max_width", 50)
    min_cap = op.get("min_width", 10)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        
        calculated_width = max(min_cap, min(max_len + padding, max_cap))
        ws.column_dimensions[col_letter].width = calculated_width

    log.append("تم ضبط عروض كافة الأعمدة آلياً لتناسب أطوال النصوص.")

def op_apply_theme(wb, ws, op, log):
    """تطبيق هوية ألوان بصرية متكاملة بضغطة واحدة"""
    theme_name = op.get("theme", "etheer_gold").lower()
    theme = THEME_PRESETS.get(theme_name, THEME_PRESETS["etheer_gold"])
    header_row = op.get("header_row") or find_real_header_row(ws)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.fill = copy.copy(theme["header_fill"])
            cell.font = copy.copy(theme["header_font"])

    if op.get("zebra_striping", False):
        for r in range(header_row + 1, ws.max_row + 1):
            if r % 2 == 0:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = copy.copy(theme["zebra_fill"])

    log.append(f"تم تطبيق النسق البصري القياسي '{theme_name}' بنجاح.")

# ==========================================
# 5. معالجات الأعمدة والصفوف والخلايا والمعادلات
# ==========================================

def op_add_column(wb, ws, op, log):
    header = op.get("header", "عمود جديد")
    after_col_name = op.get("after")
    dropdown_options = op.get("dropdown_options")
    default_value = op.get("default_value", "-")
    header_row = op.get("header_row") or find_real_header_row(ws)

    target_col = ws.max_column + 1

    if after_col_name:
        hr, hc = find_header_column(ws, after_col_name)
        if hc:
            header_row = hr
            target_col = hc + 1
        else:
            log.append(f"تنبيه: تعذر تحديد موقع '{after_col_name}'، تمت الإضافة بنهاية الجدول.")

    ws.insert_cols(target_col)

    for m in list(ws.merged_cells.ranges):
        if m.min_row < header_row and m.min_col <= target_col <= m.max_col + 1:
            ws.unmerge_cells(range_string=str(m))
            ws.merge_cells(start_row=m.min_row, start_column=m.min_col, end_row=m.max_row, end_column=m.max_col + 1)

    header_cell = ws.cell(row=header_row, column=target_col, value=header)
    ref_col = target_col - 1 if target_col > 1 else target_col + 1
    copy_cell_style(ws.cell(row=header_row, column=ref_col), header_cell)

    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=target_col, value=default_value)
        copy_cell_style(ws.cell(row=r, column=ref_col), cell)

    if dropdown_options:
        dv = DataValidation(type="list", formula1=f'"{dropdown_options}"', allow_blank=True)
        ws.add_data_validation(dv)
        col_letter = get_column_letter(target_col)
        dv.add(f"{col_letter}{header_row + 1}:{col_letter}{max(ws.max_row, header_row + 1)}")

    log.append(f"تم إضافة العمود '{header}' في الصف {header_row}، العمود {target_col}.")

def op_delete_column(wb, ws, op, log):
    col_name = op.get("header")
    if not col_name:
        raise ValueError("لم يتم تحديد اسم الهيدر للحذف.")
    hr, hc = find_header_column(ws, col_name)
    if hc:
        ws.delete_cols(hc)
        log.append(f"تم حذف العمود '{col_name}'.")
    else:
        log.append(f"تنبيه: العمود '{col_name}' غير موجود.")

def op_rename_column(wb, ws, op, log):
    old_name, new_name = op.get("old_header"), op.get("new_header")
    if not old_name or not new_name:
        raise ValueError("يجب توفير old_header و new_header.")
    hr, hc = find_header_column(ws, old_name)
    if hc:
        ws.cell(row=hr, column=hc, value=new_name)
        log.append(f"تم تعديل المسمى من '{old_name}' إلى '{new_name}'.")
    else:
        log.append(f"تنبيه: تعذر إعادة تسمية العمود '{old_name}'.")

def op_set_column_width(wb, ws, op, log):
    header, width = op.get("header"), op.get("width")
    if not header or width is None:
        raise ValueError("يجب تحديد header و width.")
    hr, hc = find_header_column(ws, header)
    if hc:
        ws.column_dimensions[get_column_letter(hc)].width = width
        log.append(f"تم تعديل عرض العمود '{header}' إلى {width}.")

def op_add_row(wb, ws, op, log):
    row_idx = op.get("row_index", ws.max_row + 1)
    ws.insert_rows(row_idx)
    log.append(f"تم إدراج صف جديد بالرقم {row_idx}.")

def op_delete_row(wb, ws, op, log):
    row_idx = op.get("row_index")
    if not row_idx:
        raise ValueError("يجب توفير row_index لحذف الصف.")
    ws.delete_rows(row_idx)
    log.append(f"تم حذف الصف رقم {row_idx}.")

def op_update_cell(wb, ws, op, log):
    addr, val = op.get("address"), op.get("value")
    if not addr:
        raise ValueError("يجب إرسال عنوان الخلية address.")
    ws[addr] = val
    log.append(f"تم تحديث الخلية {addr}.")

def op_apply_formula(wb, ws, op, log):
    """حقن معادلات إكسل الذكية المتقدمة (SUM, VLOOKUP, SUMIFS, INDEX/MATCH)"""
    addr, formula = op.get("address"), op.get("formula")
    if not addr or not formula:
        raise ValueError("يجب تحديد الخلية والمعادلة.")
    if not formula.startswith("="):
        formula = "=" + formula
    ws[addr] = formula
    log.append(f"تم تدوين المعادلة المتقدمة في الخلية {addr}.")

def op_merge_cells(wb, ws, op, log):
    range_ref = op.get("range")
    if not range_ref:
        raise ValueError("يلزم توفير نطاق range للدمج.")
    ws.merge_cells(range_ref)
    log.append(f"تم دمج النطاق {range_ref}.")

def op_unmerge_cells(wb, ws, op, log):
    range_ref = op.get("range")
    if not range_ref:
        raise ValueError("يلزم توفير نطاق range لفك الدمج.")
    ws.unmerge_cells(range_ref)
    log.append(f"تم فك دمج النطاق {range_ref}.")

# ==========================================
# 6. الشيتات والنطاقات، الرسوم والتحكم المؤسسي
# ==========================================

def op_sheet_select(wb, ws, op, log):
    sheet_name = op.get("sheet") or op.get("sheet_name")
    matched = find_sheet_flexible(wb, sheet_name)
    if not matched:
        raise ValueError(f"الشيت غير موجود: {sheet_name}")
    log.append(f"تم الانتقال إلى الشيت '{matched}'.")
    return wb[matched]

def op_sheet_create(wb, ws, op, log):
    s_name = op.get("sheet_name") or op.get("sheet") or "Sheet_New"
    matched = find_sheet_flexible(wb, s_name)
    if matched:
        log.append(f"الشيت '{matched}' قائم بالفعل وتم اختياره.")
        return wb[matched]
    new_ws = wb.create_sheet(title=s_name)
    log.append(f"تم إنشاء الشيت '{s_name}'.")
    return new_ws

def op_sheet_delete(wb, ws, op, log):
    s_name = op.get("sheet_name") or op.get("sheet")
    matched = find_sheet_flexible(wb, s_name)
    if matched:
        wb.remove(wb[matched])
        log.append(f"تم حذف الشيت '{matched}'.")
    return wb.active

def op_clear_range(wb, ws, op, log):
    range_ref = op.get("range")
    for row in ws[range_ref]:
        for cell in row:
            cell.value = None
    log.append(f"تم تفريغ النطاق {range_ref}.")

def op_color_range(wb, ws, op, log):
    range_ref, fill_color = op.get("range"), op.get("fill_color", "FFFF00")
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for row in ws[range_ref]:
        for cell in row:
            cell.fill = fill
    log.append(f"تم تلوين {range_ref} باللون {fill_color}.")

def op_border_range(wb, ws, op, log):
    range_ref = op.get("range")
    b_style, b_color = op.get("border_style", "thin"), op.get("border_color", "000000")
    border = Border(
        left=Side(style=b_style, color=b_color),
        right=Side(style=b_style, color=b_color),
        top=Side(style=b_style, color=b_color),
        bottom=Side(style=b_style, color=b_color)
    )
    for row in ws[range_ref]:
        for cell in row:
            cell.border = border
    log.append(f"تم تطبيق الحدود على النطاق {range_ref}.")

def op_conditional_formatting(wb, ws, op, log):
    """إدراج قواعد التنسيق الشرطي الديناميكي المتطور"""
    range_ref = op.get("range")
    operator = op.get("operator", "equal")
    formula_val = op.get("value", "")
    bg_color = op.get("bg_color", "FFC7CE")
    text_color = op.get("text_color", "9C0006")
    
    rule = CellIsRule(
        operator=operator,
        formula=[f'"{formula_val}"' if isinstance(formula_val, str) else str(formula_val)],
        fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
        font=Font(color=text_color, bold=True)
    )
    ws.conditional_formatting.add(range_ref, rule)
    log.append(f"تم إدراج قواعد التنسيق الشرطي الديناميكي للنطاق {range_ref}.")

def op_add_chart(wb, ws, op, log):
    """بناء الرسوم البيانية الأصلية الحية المرتبطة بالبيانات"""
    chart_type = op.get("chart_type", "bar").lower()
    title = op.get("title", "تقرير بياني مؤسسي")
    min_c, min_r = op.get("min_col", 1), op.get("min_row", 1)
    max_c, max_r = op.get("max_col", ws.max_column), op.get("max_row", ws.max_row)
    pos = op.get("cell_position", "E2")
    
    chart = LineChart() if chart_type == "line" else PieChart() if chart_type == "pie" else BarChart()
    chart.title = title
    chart.add_data(Reference(ws, min_col=min_c, min_row=min_r, max_col=max_c, max_row=max_r), titles_from_data=True)
    ws.add_chart(chart, pos)
    log.append(f"تم بناء رسم بياني أصلي حي ({chart_type}) في الموقع {pos}.")

def op_protect_sheet(wb, ws, op, log):
    """الحماية الجزئية الهرمية وقفل الخلايا الذكي"""
    password = op.get("password", "etheer2026_secure")
    ws.protection.password = password
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    log.append("تم تفعيل نظام الحماية الهرمية والقفل الذكي للشيت بنجاح.")

def op_pandas_pivot_to_sheet(wb, ws, op, log):
    s_name = find_sheet_flexible(wb, op.get("sheet") or op.get("sheet_name") or ws.title) or ws.title
    source_ws = wb[s_name]
    
    data = source_ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)

    pivot_df = pd.pivot_table(df, index=op.get("index"), values=op.get("values"), aggfunc=op.get("aggfunc", "sum"))

    target_name = op.get("target_sheet", "Pivot_Result")
    matched_target = find_sheet_flexible(wb, target_name) or target_name
    
    if matched_target in wb.sheetnames:
        target_ws = wb[matched_target]
        for row in target_ws[target_ws.dimensions]:
            for cell in row:
                cell.value = None
    else:
        target_ws = wb.create_sheet(title=matched_target)

    for r_idx, row in enumerate(pivot_df.reset_index().itertuples(index=False), start=1):
        for c_idx, val in enumerate(row, start=1):
            target_ws.cell(row=r_idx, column=c_idx, value=val)

    log.append(f"تم توليد التجميع الإحصائي Pivot بنجاح في '{target_name}'.")

def op_execute_code(wb, ws, op, log):
    code_snippet = op.get("code")
    if not code_snippet:
        raise ValueError("لم يتم إرسال كود التنفيذ البرمجي.")
    
    is_safe, error_msg = validate_python_code_security(code_snippet)
    if not is_safe:
        raise SecurityError(error_msg)

    local_scope = {
        "wb": wb, "ws": ws, "openpyxl": openpyxl, "pd": pd,
        "Font": Font, "PatternFill": PatternFill, "Alignment": Alignment,
        "Border": Border, "Side": Side, "DataValidation": DataValidation,
        "get_column_letter": get_column_letter, "column_index_from_string": column_index_from_string,
        "BarChart": BarChart, "LineChart": LineChart, "PieChart": PieChart, "Reference": Reference,
        "find_real_header_row": find_real_header_row, "find_header_column": find_header_column
    }
    exec(code_snippet, {"__builtins__": __builtins__}, local_scope)
    log.append("تم تنفيذ السكريبت البرمجي الآمن بنجاح.")

class SecurityError(Exception):
    pass

# ==========================================
# 7. توجيه العمليات ومحرك المعاملات السيادي
# ==========================================

OPERATION_MAP = {
    "execute_code": op_execute_code,
    "add_column": op_add_column,
    "delete_column": op_delete_column,
    "rename_column": op_rename_column,
    "set_column_width": op_set_column_width,
    "autofit_columns": op_autofit_columns,
    "apply_theme": op_apply_theme,
    "add_row": op_add_row,
    "delete_row": op_delete_row,
    "update_cell": op_update_cell,
    "apply_formula": op_apply_formula,
    "merge_cells": op_merge_cells,
    "unmerge_cells": op_unmerge_cells,
    "sheet_select": op_sheet_select,
    "sheet_create": op_sheet_create,
    "sheet_delete": op_sheet_delete,
    "clear_range": op_clear_range,
    "color_range": op_color_range,
    "border_range": op_border_range,
    "conditional_formatting": op_conditional_formatting,
    "add_chart": op_add_chart,
    "protect_sheet": op_protect_sheet,
    "pandas_pivot_to_sheet": op_pandas_pivot_to_sheet,
}

def execute_operations(file_path, operations):
    """محرك التنفيذ المتكامل المدعوم بحفظ الماكروز، المعاملات، والتراجع الفوري"""
    try:
        # الحفاظ التام على الحمض النووي للماكروز المؤسسية (.xlsm)
        is_macro_enabled = file_path.lower().endswith('.xlsm')
        wb = openpyxl.load_workbook(file_path, keep_vba=is_macro_enabled)
        
        # أخذ لقطة استرجاعية فائقة السرعة في الذاكرة الحية
        backup_buffer = io.BytesIO()
        wb.save(backup_buffer)

        ws = wb.active
        execution_log = []

        for idx, op in enumerate(operations):
            op_type = op.get("type")

            target_ws = get_sheet(wb, op) if op_type != "sheet_select" else ws

            if op_type == "execute_code" or ("code" in op and op_type is None):
                op_execute_code(wb, target_ws, op, execution_log)
                continue

            handler = OPERATION_MAP.get(op_type)
            if not handler:
                raise ValueError(f"العملية غير مدعومة: {op_type}")

            result = handler(wb, target_ws, op, execution_log)
            if isinstance(result, openpyxl.worksheet.worksheet.Worksheet):
                ws = result

        # فحص صحة المعادلات المدمجة والكشف عن أي تلف محتمل
        broken_formulas = scan_file_health(ws)

        wb.save(file_path)
        return {
            "success": True,
            "message": "تم تنفيذ كافة العمليات المؤسسية وحفظ التغيرات بنجاح.",
            "log": execution_log,
            "health_warnings": broken_formulas if broken_formulas else None
        }

    except Exception as e:
        return {
            "success": False,
            "error_type": type(e).__name__,
            "error": str(e),
            "rollback_status": "تم إلغاء التغييرات واستعادة الملف الأصلي فوراً في الذاكرة الحية.",
            "trace_hint": "تأكد من خلو مصفوفة الأوامر من المخالفات البرمجية أو المراجع المفقودة."
        }

# ==========================================
# 8. نقطة الدخول (CLI Entry Point)
# ==========================================

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(json.dumps({
            "success": False,
            "error": "المعلمات غير كافية. يرجى تمرير مسار الملف وسلسلة JSON."
        }, ensure_ascii=False))
        sys.exit(1)

    file_path = sys.argv[1]
    operations_json = sys.argv[2]

    try:
        operations = json.loads(operations_json)
    except json.JSONDecodeError as jde:
        print(json.dumps({
            "success": False,
            "error": f"خطأ في تحليل صيغة JSON: {str(jde)}"
        }, ensure_ascii=False))
        sys.exit(1)

    result = execute_operations(file_path, operations)
    print(json.dumps(result, ensure_ascii=False))

