import sys
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

def process_operations(file_path, operations):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # تحديد صف العناوين
        header_row = 1
        for row in range(1, min(4, ws.max_row + 1)):
            if ws.cell(row=row, column=1).value:
                header_row = row
                break

        # قاموس العناوين لتسهيل البحث
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                headers[str(val).strip()] = col

        last_added_col_letter = None

        for op in operations:
            op_type = op.get('type')

            if op_type == 'add_column':
                after = op.get('after', '')
                header = op.get('header', 'عمود جديد')
                
                target_col = ws.max_column + 1
                if after in headers:
                    target_col = headers[after] + 1
                
                ws.insert_cols(target_col)
                ws.cell(row=header_row, column=target_col, value=header)
                last_added_col_letter = get_column_letter(target_col)
                headers[header] = target_col
                
                # نسخ التنسيق بأمان تام
                source_col = target_col - 1 if target_col > 1 else target_col + 1
                if source_col <= ws.max_column:
                    for row in range(header_row, ws.max_row + 1):
                        source_cell = ws.cell(row=row, column=source_col)
                        target_cell = ws.cell(row=row, column=target_col)
                        if source_cell.has_style:
                            if source_cell.font: target_cell.font = copy(source_cell.font)
                            if source_cell.border: target_cell.border = copy(source_cell.border)
                            if source_cell.fill: target_cell.fill = copy(source_cell.fill)
                            if source_cell.number_format: target_cell.number_format = copy(source_cell.number_format)
                            if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)

            elif op_type == 'add_validation':
                formulae = op.get('formulae', '')
                address = op.get('address', '')
                
                dv = DataValidation(type="list", formula1=f'"{formulae}"', allow_blank=True)
                ws.add_data_validation(dv)
                
                if address:
                    dv.add(address)
                elif last_added_col_letter:
                    dv.add(f"{last_added_col_letter}{header_row + 1}:{last_added_col_letter}{ws.max_row}")
                else:
                    col_letter = get_column_letter(ws.max_column)
                    dv.add(f"{col_letter}{header_row + 1}:{col_letter}{ws.max_row}")

            elif op_type == 'autofit_columns':
                for col in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)
        return {"success": True, "message": "تم التنفيذ بنجاح"}

    except Exception as e:
        return {"success": False, "error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(json.dumps({"success": False, "error": "Missing arguments"}))
        sys.exit(1)
        
    file_path = sys.argv[1]
    try:
        operations = json.loads(sys.argv[2])
        result = process_operations(file_path, operations)
        print(json.dumps(result))
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Execution failed: {str(e)}"}))
