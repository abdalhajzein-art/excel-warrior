"""
api/core/excel_preview.py – Sovereign Excel Engine (Strict Tabular Edition)
⚡ محرك إكسل سيادي بمعمارية قياسية صارمة: ترويسة من الصف الأول، دعم RTL، تحليل مخطط دقيق، وتوليد نظيف.
✅ تمت إضافة metadata متوافقة مع file_fingerprint.js
✅ تم إصلاح اقتطاع النص وتحسين خوارزمية معاينة البيانات لضمان الرؤية المطلقة لـ Gemini
"""

import sys
import json
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border

MAX_PREVIEW_ROWS = 25
MAX_FORMULAS = 20

def safe_str(value):
    return "" if value is None else str(value).strip()

def find_header_row_and_schema(ws, max_rows=10):
    """
    خوارزمية ذكية لاكتشاف صف الترويسة واستخراج مخطط الأعمدة بدقة.
    في البناء القياسي الصحيح، تبدأ الترويسة دائماً من الصف الأول.
    """
    best_row_idx = 1
    max_score = -1
    schema = {}

    if ws.max_row == 0 or ws.max_column == 0:
        return best_row_idx, schema

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        score = 0
        current_schema = {}
        for col_idx, val in enumerate(row, start=1):
            if val is not None:
                sval = str(val).strip()
                if sval:
                    # 🛡️ استبعاد خلايا المعادلات لضمان عدم الخلط بين العناوين والبيانات
                    if sval.startswith('='):
                        score -= 5
                    else:
                        score += 1
                        current_schema[col_idx] = sval
        
        if score > max_score:
            max_score = score
            best_row_idx = r_idx
            schema = current_schema
            
    return best_row_idx, schema

def extract_sheet_preview(wb, sheet_name, max_rows=MAX_PREVIEW_ROWS):
    ws = wb[sheet_name]
    header_row_idx, schema = find_header_row_and_schema(ws)

    preview_rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        preview_rows.append([safe_str(c) for c in row])

    merged_cells = [str(rng) for rng in list(ws.merged_cells.ranges)[:20]]

    return {
        "sheet": sheet_name,
        "rows_count": ws.max_row,
        "columns_count": ws.max_column,
        "detected_header_row": header_row_idx,
        "columns_schema": schema, 
        "preview_rows": preview_rows,
        "merged": merged_cells
    }

def generate_new_excel_file(output_path, sheet_name, headers, rows_data):
    """
    توليد ملف إكسل جديد كلياً من الصفر بمعمارية قياسية صحيحة:
    - ترويسة تبدأ فوراً من الصف الأول (Row 1).
    - اتجاه ورقة العمل من اليمين لليسار (RTL) لدعم البيئة العربية.
    - تجميد صف الترويسة (Freeze Panes) لتسهيل التمرير.
    - تنسيق موحد واحترافي لجميع الخلايا.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 🌐 ضبط اتجاه ورقة العمل إلى اليمين لليسار (RTL)
    ws.sheet_view.rightToLeft = True

    # 1. إعداد ترويسة الجدول في الصف الأول حصراً
    header_row_idx = 1
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=openpyxl.styles.borders.Side(style='thin', color='D9D9D9'),
        right=openpyxl.styles.borders.Side(style='thin', color='D9D9D9'),
        top=openpyxl.styles.borders.Side(style='thin', color='D9D9D9'),
        bottom=openpyxl.styles.borders.Side(style='thin', color='D9D9D9')
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    ws.row_dimensions[header_row_idx].height = 26

    # 2. إدخال بيانات الصفوف ابتداءً من الصف الثاني (Row 2)
    data_font = Font(name="Calibri", size=11)
    for r_offset, row_data in enumerate(rows_data):
        current_row = header_row_idx + 1 + r_offset
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 20

    # 3. تجميد صف الترويسة العلوي لضمان بقائه أثناء التمرير
    ws.freeze_panes = "A2"

    # 4. ضبط عرض الأعمدة تلقائياً لمنع اقتطاع أي نص
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    wb.save(output_path)
    wb.close()
    return {"status": "success", "output": output_path, "message": "تم توليد الملف بنجاح بمعمارية قياسية نظيفة (ترويسة من الصف الأول و RTL)."}

def execute_excel_operation(file_path, output_path, operation_type, target_keyword=None, options=None):
    """
    محرك تنفيذ التعديلات (إضافة أعمدة، قوائم منسدلة، إلخ) مع الاعتماد على الكشف الديناميكي للترويسة.
    """
    options = options or {}
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row_idx, schema = find_header_row_and_schema(ws)

    if operation_type == "add_column_with_dropdown":
        col_idx = None
        for c_idx, h_val in schema.items():
            if target_keyword and target_keyword in h_val:
                col_idx = c_idx
                break

        if not col_idx:
            raise ValueError(f"لم يتم العثور على عمود مطابق لـ '{target_keyword}' في الترويسة المكتشفة.")

        target_col_idx = col_idx + 1
        ws.insert_cols(target_col_idx)

        new_header_val = options.get("new_header", "ملاحظات")
        header_cell = ws.cell(row=header_row_idx, column=target_col_idx, value=new_header_val)
        ref_header = ws.cell(row=header_row_idx, column=col_idx)
        
        header_cell.font = Font(name=ref_header.font.name, size=ref_header.font.size, bold=True, color=ref_header.font.color)
        if ref_header.fill:
            header_cell.fill = PatternFill(start_color=ref_header.fill.start_color, end_color=ref_header.fill.end_color, fill_type=ref_header.fill.fill_type)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = Border(
            left=ref_header.border.left, right=ref_header.border.right,
            top=ref_header.border.top, bottom=ref_header.border.bottom
        )

        dropdown_values = options.get("dropdown_values", [])
        dv = None
        if dropdown_values:
            reasons_str = '"' + ','.join(dropdown_values) + '"'
            dv = DataValidation(type="list", formula1=reasons_str, allow_blank=True)
            ws.add_data_validation(dv)

        for row in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=target_col_idx)
            ref_data = ws.cell(row=row, column=col_idx)
            
            cell.font = Font(name=ref_data.font.name, size=ref_data.font.size, color=ref_data.font.color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=ref_data.border.left, right=ref_data.border.right,
                top=ref_data.border.top, bottom=ref_data.border.bottom
            )
            if dv:
                dv.add(cell)

    wb.save(output_path)
    wb.close()
    return {"status": "success", "output": output_path, "message": "تم تنفيذ العملية بنجاح."}

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "لم يتم تمرير مسار ملف أو معاملات تشغيل."}, ensure_ascii=False))
        return

    arg1 = sys.argv[1]

    # 🎯 الحالة الأولى: معاملات JSON (التوليد الجديد أو التعديلات)
    if arg1.startswith("{") or len(sys.argv) >= 3:
        try:
            op_data = json.loads(arg1) if arg1.startswith("{") else json.loads(sys.argv[2])
            operation_type = op_data.get("operation_type")
            output_path = op_data.get("output_path", "generated_report.xlsx")
            options = op_data.get("options", {})
            
            if operation_type == "generate_new":
                headers = options.get("headers", [])
                rows_data = options.get("rows_data", [])
                sheet_name = options.get("sheet_name", "التقرير")
                result = generate_new_excel_file(output_path, sheet_name, headers, rows_data)
            else:
                file_path = arg1 if not arg1.startswith("{") else op_data.get("file_path")
                result = execute_excel_operation(file_path, output_path, operation_type, op_data.get("target_keyword"), options)
                
            print(json.dumps(result, ensure_ascii=False))
        except Exception as e:
            print(json.dumps({"error": str(e)}, ensure_ascii=False))
        return

    # 🔍 الحالة الثانية: معاينة الملف واستخراج المخطط (مع دعم metadata)
    file_path = arg1
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        sheets = wb.sheetnames
        previews = [extract_sheet_preview(wb, sheet) for sheet in sheets]
        
        # ✅ بناء metadata متوافق مع file_fingerprint.js
        metadata = {
            "sheets": sheets,
            "rowCounts": {},
            "columnCounts": {}
        }
        
        for p in previews:
            sheet_name = p["sheet"]
            metadata["rowCounts"][sheet_name] = p.get("rows_count", 0)
            metadata["columnCounts"][sheet_name] = p.get("columns_count", 0)
        
        # ✅ بناء النص الموجز للمعاينة (النسخة السيادية المصححة)
        preview_text_parts = []
        for p in previews[:2]:
            sheet_name = p["sheet"]
            
            # 1. تصفية الأسطر الفارغة تماماً (لتجاوز الترويسات والصفوف البيضاء)
            valid_rows = [row for row in p.get("preview_rows", []) if any(str(cell).strip() for cell in row)]
            
            # 2. أخذ أول 10 أسطر فعلية مليئة بالبيانات بدلاً من 3
            rows_to_show = valid_rows[:10]
            
            if rows_to_show:
                # 3. تحويلها إلى JSON مقروء وإلغاء حد الـ 200 حرف القاتل (رفع الحد إلى 4000 حرف)
                rows_str = json.dumps(rows_to_show, ensure_ascii=False)
                preview_text_parts.append(f"[{sheet_name}]: {rows_str[:4000]}")
        
        output = {
            "file": file_path,
            "sheets_count": len(sheets),
            "sheets": sheets,
            "previews": previews,
            "metadata": metadata,
            "rows": sum(metadata["rowCounts"].values()),
            "columns": max(metadata["columnCounts"].values()) if metadata["columnCounts"] else 0,
            "text": "\n".join(preview_text_parts) if preview_text_parts else "تم استخراج معاينة الملف"
        }
        
        print(json.dumps(output, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass

if __name__ == "__main__":
    main()
