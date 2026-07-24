import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import traceback

class AlatheerUltimateEngine:
    """محرك الكواليس الخارق والسيادي - الإصدار الهندسي المتقدم (الأثير)"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        self.header_row = 1
        self.headers = []
    
    def load_file(self, file_path, sheet_name=None):
        try:
            # تحميل الملف مع الحفاظ التام على المعادلات والصيغ الأصلية
            self.wb = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as e:
            raise Exception(f"فشل تحميل الملف برمجياً: {str(e)}")
            
        if sheet_name and sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.active
            
        self._smart_detect_structure()

    def _smart_detect_structure(self):
        """كشف هندسي ذكي جداً لصف العناوين بغض النظر عن تصميم الملف"""
        max_r = min(self.ws.max_row + 1, 25)
        max_c = min(self.ws.max_column + 1, 30)
        
        best_row = 1
        max_non_empty = 0
        
        for r in range(1, max_r):
            count = sum(1 for c in range(1, max_c) if self.ws.cell(row=r, column=c).value is not None)
            if count > max_non_empty:
                max_non_empty = count
                best_row = r
                
        self.header_row = best_row
        
        # استخراج الأعمدة بذكاء وتنظيف الأسماء
        self.headers = []
        for c in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=self.header_row, column=c).value
            if val is not None and str(val).strip() != "":
                self.headers.append(str(val).strip())
            else:
                self.headers.append(f"Col_{c}")

    def intelligent_read_and_analyze(self):
        """وحدة الاستشعار والتحليل العميق: تجميع البيانات كاملة للملف دون إهدار للتوكنز"""
        rows_data = []
        full_records = []
        
        # قراءة كل الصفوف البيانية المتاحة حتى لو كانت مئات الصفوف
        for r in range(self.header_row + 1, self.ws.max_row + 1):
            row_dict = {}
            has_val = False
            for c, h in enumerate(self.headers, start=1):
                cell_val = self.ws.cell(row=r, column=c).value
                if cell_val is not None:
                    has_val = True
                    row_dict[h] = cell_val
            if has_val:
                rows_data.append(row_dict)
                full_records.append({"row_index": r, "data": row_dict})

        return {
            "success": True,
            "is_read_only": True,
            "action_type": "read",
            "metadata": {
                "total_rows": len(rows_data),
                "total_columns": self.ws.max_column,
                "header_row": self.header_row,
                "headers": self.headers,
                "sample_data": rows_data[:50], # إرسال عينة وافية لعقل الذكاء الاصطناعي
                "full_records": full_records   # البيانات الكاملة للتحليل الدقيق داخل بايثون إن احتجنا
            },
            "message": "✅ تم استقراء وتحليل بيانات الملف بذكاء هندسي تام."
        }

    def execute_advanced_operations(self, plan, instruction):
        """وحدة التعديل والحسابات الذكية (تلوين، إضافة أعمدة، أو استخراج إحصائيات دقيقة)"""
        instruction_lower = instruction.strip().lower()
        
        # 1. تنفيذ خطة الأعمدة الجديدة إن وجدت في الطلب
        new_columns = plan.get("newColumns", [])
        for col in new_columns:
            col_name = col.get("name", "عمود جديد")
            formula = col.get("formula", None)
            
            insert_pos = self.ws.max_column + 1
            self.ws.insert_cols(insert_pos, amount=1)
            
            header_cell = self.ws.cell(row=self.header_row, column=insert_pos, value=col_name)
            self._apply_style(header_cell, 'header')
            
            for r in range(self.header_row + 1, self.ws.max_row + 1):
                if self.ws.cell(row=r, column=1).value is not None:
                    cell = self.ws.cell(row=r, column=insert_pos)
                    self._apply_style(cell, 'data')
                    if formula:
                        cell.value = formula.replace("{row}", str(r))
                    else:
                        cell.value = 0

        # 2. التلوين الشرطي والتمييز الذكي بناءً على الأوامر اللفظية
        if "لون" in instruction_lower or "تمييز" in instruction_lower or "تنبيه" in instruction_lower:
            accent_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            for r in range(self.header_row + 1, self.ws.max_row + 1):
                for c in range(1, self.ws.max_column + 1):
                    val = str(self.ws.cell(row=r, column=c).value or "").lower()
                    if any(w in instruction_lower for w in ["غياب", "تأخير", "مطلوب صيانة", "خطر"]):
                        if "غياب" in val or "تأخير" in val or "مطلوب" in val:
                            self.ws.cell(row=r, column=c).fill = accent_fill

        return {
            "success": True,
            "is_read_only": False,
            "action_type": "modify",
            "message": "✅ تم تنفيذ العمليات الهندسية والتعديلات على الملف بنجاح تام."
        }

    def _apply_style(self, cell, style_type='data'):
        if style_type == 'header':
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        else:
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
        )

def main():
    try:
        raw_input = sys.stdin.read()
        if not raw_input:
            print(json.dumps({"success": False, "error": "لم يتم استلام بيانات في المحرك"}), flush=True)
            return
            
        input_data = json.loads(raw_input)
        input_path = input_data.get("inputPath")
        output_path = input_data.get("outputPath")
        action = input_data.get("action", "read")
        plan = input_data.get("plan", {})
        instruction = plan.get("summary", input_data.get("instruction", ""))
        sheet_name = input_data.get("sheetName")
        
        engine = AlatheerUltimateEngine()
        engine.load_file(input_path, sheet_name)
        
        # توجيه ذكي حسب طلب المستخدم
        if action == "formulas" or "دوال" in instruction.lower() or "معادلات" in instruction.lower():
            # استخراج المعادلات
            formulas_found = []
            for r in range(1, engine.ws.max_row + 1):
                for c in range(1, engine.ws.max_column + 1):
                    val = str(engine.ws.cell(row=r, column=c).value or "")
                    if val.startswith("="):
                        formulas_found.append(f"الخلية {get_column_letter(c)}{r}: {val}")
            result = {
                "success": True,
                "is_read_only": True,
                "action_type": "formulas",
                "formulas_list": formulas_found,
                "message": f"تم رصد {len(formulas_found)} معادلة."
            }
        elif action == "read" or any(w in instruction.lower() for w in ["اقرأ", "عرض", "محتوى", "ملخص", "استعرض", "حلل", "من", "أعلى"]):
            result = engine.intelligent_read_and_analyze()
        else:
            result = engine.execute_advanced_operations(plan, instruction)
            
        # حفظ الملف الناتج إذا كان هناك تعديل
        if not result.get("is_read_only", False) and output_path:
            engine.wb.save(output_path)
            
        print(json.dumps(result), flush=True)
        
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e), "traceback": traceback.format_exc()}), flush=True)

if __name__ == "__main__":
    main()

