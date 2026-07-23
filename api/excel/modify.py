import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def modify_excel():
    try:
        input_data = json.loads(sys.stdin.read())
        file_path = input_data.get("filePath")
        new_columns = input_data.get("newColumns", [])
        target_column = input_data.get("targetColumn")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 1. البحث عن صف العناوين ديناميكياً
        header_row = 1
        for r in range(1, min(ws.max_row + 1, 10)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or "").strip()
                if "رقم الموظف" in val or "اسم الموظف" in val or "الغياب" in val or "اليوم" in val:
                    header_row = r
                    break
            if header_row != 1:
                break

        # 2. تحديد موقع الإدراج
        target_col_idx = None
        for c in range(1, ws.max_column + 1):
            h_val = str(ws.cell(row=header_row, column=c).value or "").strip()
            if target_column and target_column in h_val:
                target_col_idx = c
                break
            elif not target_column and ("الغياب" in h_val or h_val == "غياب"):
                target_col_idx = c
                break

        insert_pos = (target_col_idx + 1) if target_col_idx else (ws.max_column + 1)

        # 3. سحب قالب التنسيق من العمود المجاور (قبل الإدراج لضمان دقة الألوان والخطوط)
        sample_col = target_col_idx if target_col_idx else (insert_pos - 1)
        sample_header_cell = ws.cell(row=header_row, column=sample_col)
        
        # التقاط التنسيقات الأصلية لرأس العمود
        ref_font = sample_header_cell.font
        ref_fill = sample_header_cell.fill
        ref_alignment = sample_header_cell.alignment
        ref_border = sample_header_cell.border

        # التقاط تنسيق خلية عادية صفية
        sample_data_cell = ws.cell(row=header_row + 1, column=sample_col)
        ref_data_font = sample_data_cell.font
        ref_data_alignment = sample_data_cell.alignment
        ref_data_border = sample_data_cell.border

        # 4. إدراج الأعمدة الجديدة
        num_cols = len(new_columns)
        ws.insert_cols(insert_pos, amount=num_cols)

        # 5. تعبئة العناوين الجديدة مع نسخ الألوان والتنسيق الأصلي حرفياً
        for idx, col_name in enumerate(new_columns):
            current_col = insert_pos + idx
            
            # تطبيق تنسيق رأس العمود الأصلي
            header_cell = ws.cell(row=header_row, column=current_col, value=col_name)
            if ref_font: header_cell.font = Font(name=ref_font.name, size=ref_font.size, bold=True, color=ref_font.color)
            if ref_fill: header_cell.fill = PatternFill(fill_type=ref_fill.fill_type, start_color=ref_fill.start_color, end_color=ref_fill.end_color)
            if ref_alignment: header_cell.alignment = Alignment(horizontal=ref_alignment.horizontal, vertical=ref_alignment.vertical)
            if ref_border: header_cell.border = ref_border

            # تعبئة الصفوف وتطبيق التنسيق القياسي للبيانات
            for r in range(header_row + 1, ws.max_row + 1):
                if ws.cell(row=r, column=1).value is not None:
                    cell = ws.cell(row=r, column=current_col)
                    if "سبب" in col_name:
                        cell.value = "مرض"
                    elif "ملاحظات" in col_name:
                        cell.value = "بدون ملاحظات"
                    else:
                        cell.value = "-"
                    
                    if ref_data_font: cell.font = ref_data_font
                    if ref_data_alignment: cell.alignment = ref_data_alignment
                    if ref_data_border: cell.border = ref_data_border

        wb.save(file_path)
        print(json.dumps({"success": True, "message": "تم إدراج الأعمدة مع نسخ الألوان والتنسيقات الأصلية بنجاح"}))

    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    modify_excel()
